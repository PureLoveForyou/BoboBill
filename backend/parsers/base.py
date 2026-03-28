import re
import csv
import io
from datetime import datetime
from typing import List, Dict, Optional
from openpyxl import load_workbook


def decode_content(content: bytes) -> str:
    """Try multiple encodings to decode file content."""
    for enc in ['utf-8', 'utf-8-sig', 'gbk', 'gb18030']:
        try:
            return content.decode(enc)
        except (UnicodeDecodeError, LookupError):
            continue
    raise ValueError("无法解码文件内容")


def parse_date(date_str: str) -> str:
    """Parse date string to YYYY-MM-DD."""
    if not date_str:
        return ""
    date_str = date_str.strip()
    try:
        return datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d")
    except ValueError:
        return date_str[:10] if len(date_str) >= 10 else date_str


def parse_amount(amount_str: str) -> Optional[float]:
    """Clean currency symbols and parse amount."""
    cleaned = re.sub(r'[¥￥,\s]', '', amount_str.strip())
    try:
        return float(cleaned)
    except ValueError:
        return None


def determine_type(direction: str) -> str:
    """Determine income/expense from direction text."""
    return 'income' if '收入' in direction else 'expense'


def categorize_transaction(name: str, trans_type: str) -> str:
    """Categorize a transaction based on merchant name and type."""
    name_lower = (name or "").lower()
    categories = {
        '餐饮': ['美团', '饿了么', '外卖', '餐厅', '食品', '星巴克', '肯德基', '麦当劳', '奶茶', '咖啡', '饭店', '小吃', '超市', '便利店'],
        '交通': ['滴滴', '打车', '地铁', '公交', '加油', '停车', '出行', '火车', '机票', '高铁'],
        '购物': ['淘宝', '京东', '拼多多', '天猫', '购物', '商城', '网购'],
        '娱乐': ['游戏', '视频', '音乐', '电影', '会员', '充值'],
        '医疗': ['医院', '药店', '医疗', '健康'],
        '教育': ['教育', '培训', '课程', '学习'],
        '转账': ['转账', '红包', '收款'],
        '工资': ['工资', '薪资', '代发'],
        '投资': ['理财', '基金', '股票', '收益'],
    }
    for cat, keywords in categories.items():
        for kw in keywords:
            if kw in name_lower:
                return cat
    if '转账' in trans_type or '红包' in trans_type:
        return '转账'
    return '其他'


def build_col_map(headers: List[str]) -> Dict[str, int]:
    """Build column index mapping from header strings."""
    col_map = {}
    for i, h in enumerate(headers):
        h_s = h.strip()
        if ('交易类型' in h_s or '交易分类' in h_s) and '时间' not in h_s:
            col_map['type'] = i
        elif '交易时间' in h_s:
            col_map['date'] = i
        elif '交易对方' in h_s:
            col_map['merchant'] = i
        elif '商品说明' in h_s or '商品' in h_s:
            col_map['name'] = i
        elif '收/支' in h_s:
            col_map['direction'] = i
        elif '金额(元)' in h_s or '金额' in h_s:
            col_map['amount'] = i
        elif '收/付款方式' in h_s or '支付方式' in h_s:
            col_map['payment'] = i
        elif '交易单号' in h_s or '交易订单号' in h_s or '交易号' in h_s:
            col_map['transaction_id'] = i
        elif '备注' in h_s:
            col_map['note'] = i
    return col_map


def row_to_bill(row_values: List[str], col_map: Dict[str, int], platform: str) -> Optional[dict]:
    """Convert a single row of values into a bill dict."""
    if not row_values or len(row_values) < 5:
        return None

    def get_val(key):
        return row_values[col_map[key]].strip() if key in col_map and col_map[key] < len(row_values) else ""

    date_str = get_val('date')
    if not date_str:
        return None

    amount = parse_amount(get_val('amount'))
    if amount is None:
        return None

    bill_type = determine_type(get_val('direction'))
    amount = abs(amount) if bill_type == 'income' else -abs(amount)

    merchant = get_val('merchant')
    name = get_val('name') or merchant
    note = get_val('note')

    return {
        'name': (name[:100] if name else "未知交易"),
        'amount': amount,
        'type': bill_type,
        'date': parse_date(date_str),
        'category': categorize_transaction(name, get_val('type')),
        'platform': platform,
        'merchant': (merchant[:100] if merchant else ""),
        'note': (note[:200] if note else ""),
        'transaction_id': get_val('transaction_id'),
    }


def find_csv_header(lines: list, checker) -> int:
    """Find CSV header line index by checker function."""
    for i, line in enumerate(lines):
        if checker(line):
            return i
    return -1


def find_excel_header(ws, checker):
    """Find Excel header row. Returns (row_index, headers) or (None, None)."""
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        vals = [str(c) if c is not None else '' for c in row]
        if checker(vals):
            return row_idx, vals
    return None, None


def parse_csv_rows(rows, col_map: Dict[str, int], platform: str) -> list:
    """Iterate CSV rows and convert to bill dicts."""
    bills = []
    for row in rows:
        if not row:
            continue
        if '---' in str(row[0]) or '以上是' in str(row[0]):
            break
        bill = row_to_bill(row, col_map, platform)
        if bill:
            bills.append(bill)
    return bills


def parse_excel_rows(ws, header_row_idx: int, col_map: Dict[str, int], platform: str) -> list:
    """Iterate Excel rows and convert to bill dicts."""
    bills = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        row_values = [str(c) if c is not None else '' for c in row]
        if '---' in row_values[0] or '以上是' in row_values[0]:
            break
        bill = row_to_bill(row_values, col_map, platform)
        if bill:
            bills.append(bill)
    return bills


def detect_platform(content: bytes, filename: str) -> str:
    """Detect bill platform (wechat/alipay) from file content."""
    is_excel = filename.lower().endswith(('.xlsx', '.xls'))
    wechat_kw = ['微信支付', '微信昵称', '微信支付账单明细', '商户消费', '零钱', '微信红包']
    alipay_kw = ['支付宝', '交易分类', '支付宝（中国）网络技术有限公司', '收/付款方式', '商品说明']

    if is_excel:
        try:
            wb = load_workbook(io.BytesIO(content))
            ws = wb.active
            text = ' '.join(
                ' '.join(str(c) if c else '' for c in row)
                for row in ws.iter_rows(min_row=1, max_row=30, values_only=True)
            )
        except Exception:
            return 'unknown'
    else:
        try:
            text = decode_content(content)
        except Exception:
            return 'unknown'

    wechat_score = sum(text.count(k) for k in wechat_kw)
    alipay_score = sum(text.count(k) for k in alipay_kw)

    if wechat_score > alipay_score and wechat_score > 0:
        return 'wechat'
    if alipay_score > wechat_score and alipay_score > 0:
        return 'alipay'
    return 'unknown'
