from fastapi.middleware.cors import CORSMiddleware
from fastapi import FastAPI, HTTPException, UploadFile, File, Form
from pydantic import BaseModel
from typing import List, Optional
from tinydb import TinyDB, Query
import os
import csv
import io
import re
from datetime import datetime
from openpyxl import load_workbook

if not os.path.exists("db.json"):
    open("db.json", "w").close()

db = TinyDB("db.json")
Bill = Query()

app = FastAPI(
    title="BoboBill API",
    description="智能账单管理助手",
    version="0.2.0"
)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

class BillModel(BaseModel):
    id: Optional[int] = None
    name: str
    amount: float
    type: str
    date: str
    category: Optional[str] = None
    platform: Optional[str] = None
    merchant: Optional[str] = None
    note: Optional[str] = None
    transaction_id: Optional[str] = None

class ImportResult(BaseModel):
    success: int
    skipped: int
    total: int
    message: str

@app.get("/bills", response_model=List[BillModel])
def get_bills(
    category: Optional[str] = None,
    platform: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    bills = db.all()
    result = []
    for bill in bills:
        bill["id"] = bill.doc_id
        
        if category and bill.get("category") != category:
            continue
        if platform and bill.get("platform") != platform:
            continue
        if start_date and bill.get("date", "") < start_date:
            continue
        if end_date and bill.get("date", "") > end_date:
            continue
        
        result.append(bill)
    
    result.sort(key=lambda x: x.get("date", ""), reverse=True)
    return result

@app.post("/bills", response_model=BillModel)
def add_bill(bill: BillModel):
    bill_dict = bill.model_dump()
    bill_dict.pop("id", None)
    doc_id = db.insert(bill_dict)
    bill_dict["id"] = doc_id
    return bill_dict

@app.delete("/bills/{bill_id}")
def delete_bill(bill_id: int):
    if db.remove(doc_ids=[bill_id]):
        return {"message": "删除成功"}
    else:
        raise HTTPException(status_code=404, detail="账单未找到")

@app.delete("/bills")
def clear_all_bills():
    db.truncate()
    return {"message": "所有账单已清空"}

def parse_wechat_csv(content: str) -> List[dict]:
    bills = []
    lines = content.strip().split('\n')
    
    header_idx = -1
    for i, line in enumerate(lines):
        if '交易时间' in line and '交易类型' in line:
            header_idx = i
            break
    
    if header_idx == -1:
        raise ValueError("无法识别微信账单格式：找不到表头")
    
    delimiter = '\t' if '\t' in lines[header_idx] else ','
    reader = csv.reader(lines[header_idx:], delimiter=delimiter)
    headers = next(reader)
    headers = [h.strip() for h in headers]
    
    col_map = {}
    for i, h in enumerate(headers):
        if '交易时间' in h:
            col_map['date'] = i
        elif '交易类型' in h:
            col_map['type'] = i
        elif '交易对方' in h:
            col_map['merchant'] = i
        elif '商品' in h:
            col_map['name'] = i
        elif '收/支' in h:
            col_map['direction'] = i
        elif '金额(元)' in h or '金额' in h:
            col_map['amount'] = i
        elif '支付方式' in h:
            col_map['payment'] = i
        elif '交易单号' in h:
            col_map['transaction_id'] = i
        elif '备注' in h:
            col_map['note'] = i
    
    for row in reader:
        if not row or len(row) < 5:
            continue
        
        if '---' in row[0] or '以上是' in row[0]:
            break
        
        try:
            get_val = lambda key: row[col_map[key]].strip() if key in col_map and col_map[key] < len(row) else ""
            
            date_str = get_val('date')
            if not date_str:
                continue
            
            try:
                dt = datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
                date_formatted = dt.strftime("%Y-%m-%d")
            except:
                date_formatted = date_str[:10] if len(date_str) >= 10 else date_str
            
            direction = get_val('direction')
            bill_type = 'income' if '收入' in direction else 'expense' if '支出' in direction else 'expense'
            
            amount_str = get_val('amount')
            amount_str = re.sub(r'[¥￥,\s]', '', amount_str)
            try:
                amount = float(amount_str)
            except:
                continue
            
            if bill_type == 'expense':
                amount = -abs(amount)
            else:
                amount = abs(amount)
            
            merchant = get_val('merchant')
            name = get_val('name') or merchant
            
            category = categorize_transaction(name, get_val('type'))
            
            bill = {
                'name': name[:100] if name else "未知交易",
                'amount': amount,
                'type': bill_type,
                'date': date_formatted,
                'category': category,
                'platform': 'wechat',
                'merchant': merchant[:100] if merchant else "",
                'note': get_val('note')[:200] if get_val('note') else "",
                'transaction_id': get_val('transaction_id')
            }
            
            bills.append(bill)
            
        except Exception as e:
            print(f"解析行失败: {row}, 错误: {e}")
            continue
    
    return bills

def parse_wechat_excel(content: bytes) -> List[dict]:
    bills = []
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    
    header_idx = None
    headers = []
    
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        row_values = [str(cell) if cell is not None else '' for cell in row]
        if '交易时间' in str(row_values) and '交易类型' in str(row_values):
            header_idx = row_idx
            headers = row_values
            break
    
    if header_idx is None:
        raise ValueError("无法识别微信账单格式：找不到表头")
    
    col_map = {}
    for i, h in enumerate(headers):
        h_str = str(h).strip()
        if '交易时间' in h_str:
            col_map['date'] = i
        elif '交易类型' in h_str:
            col_map['type'] = i
        elif '交易对方' in h_str:
            col_map['merchant'] = i
        elif '商品' in h_str:
            col_map['name'] = i
        elif '收/支' in h_str:
            col_map['direction'] = i
        elif '金额(元)' in h_str or '金额' in h_str:
            col_map['amount'] = i
        elif '支付方式' in h_str:
            col_map['payment'] = i
        elif '交易单号' in h_str:
            col_map['transaction_id'] = i
        elif '备注' in h_str:
            col_map['note'] = i
    
    for row in ws.iter_rows(min_row=header_idx + 1, values_only=True):
        row_values = [str(cell) if cell is not None else '' for cell in row]
        
        if not row_values or len(row_values) < 5:
            continue
        
        if '---' in row_values[0] or '以上是' in row_values[0]:
            break
        
        try:
            get_val = lambda key: row_values[col_map[key]].strip() if key in col_map and col_map[key] < len(row_values) else ""
            
            date_str = get_val('date')
            if not date_str:
                continue
            
            try:
                dt = datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
                date_formatted = dt.strftime("%Y-%m-%d")
            except:
                date_formatted = date_str[:10] if len(date_str) >= 10 else date_str
            
            direction = get_val('direction')
            bill_type = 'income' if '收入' in direction else 'expense' if '支出' in direction else 'expense'
            
            amount_str = get_val('amount')
            amount_str = re.sub(r'[¥￥,\s]', '', amount_str)
            try:
                amount = float(amount_str)
            except:
                continue
            
            if bill_type == 'expense':
                amount = -abs(amount)
            else:
                amount = abs(amount)
            
            merchant = get_val('merchant')
            name = get_val('name') or merchant
            
            category = categorize_transaction(name, get_val('type'))
            
            bill = {
                'name': name[:100] if name else "未知交易",
                'amount': amount,
                'type': bill_type,
                'date': date_formatted,
                'category': category,
                'platform': 'wechat',
                'merchant': merchant[:100] if merchant else "",
                'note': get_val('note')[:200] if get_val('note') else "",
                'transaction_id': get_val('transaction_id')
            }
            
            bills.append(bill)
            
        except Exception as e:
            print(f"解析行失败: {row_values}, 错误: {e}")
            continue
    
    return bills

def parse_alipay_csv(content: str) -> List[dict]:
    bills = []
    lines = content.strip().split('\n')
    
    header_idx = -1
    for i, line in enumerate(lines):
        if '交易时间' in line and ('交易分类' in line or '交易对方' in line):
            header_idx = i
            break
    
    if header_idx == -1:
        raise ValueError("无法识别支付宝账单格式：找不到表头")
    
    delimiter = '\t' if '\t' in lines[header_idx] else ','
    reader = csv.reader(lines[header_idx:], delimiter=delimiter)
    headers = next(reader)
    headers = [h.strip() for h in headers]
    
    col_map = {}
    for i, h in enumerate(headers):
        if '交易时间' in h:
            col_map['date'] = i
        elif '交易分类' in h or '交易类型' in h:
            col_map['type'] = i
        elif '交易对方' in h:
            col_map['merchant'] = i
        elif '商品说明' in h or '商品' in h:
            col_map['name'] = i
        elif '收/支' in h:
            col_map['direction'] = i
        elif '金额' in h:
            col_map['amount'] = i
        elif '收/付款方式' in h or '支付方式' in h:
            col_map['payment'] = i
        elif '交易订单号' in h or '交易号' in h:
            col_map['transaction_id'] = i
        elif '备注' in h:
            col_map['note'] = i
    
    for row in reader:
        if not row or len(row) < 5:
            continue
        
        if '---' in row[0] or '以上是' in row[0]:
            break
        
        try:
            get_val = lambda key: row[col_map[key]].strip() if key in col_map and col_map[key] < len(row) else ""
            
            date_str = get_val('date')
            if not date_str:
                continue
            
            try:
                dt = datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
                date_formatted = dt.strftime("%Y-%m-%d")
            except:
                date_formatted = date_str[:10] if len(date_str) >= 10 else date_str
            
            direction = get_val('direction')
            bill_type = 'income' if '收入' in direction else 'expense' if '支出' in direction else 'expense'
            
            amount_str = get_val('amount')
            amount_str = re.sub(r'[¥￥,\s]', '', amount_str)
            try:
                amount = float(amount_str)
            except:
                continue
            
            if bill_type == 'expense':
                amount = -abs(amount)
            else:
                amount = abs(amount)
            
            merchant = get_val('merchant')
            name = get_val('name') or merchant
            
            category = categorize_transaction(name, get_val('type'))
            
            bill = {
                'name': name[:100] if name else "未知交易",
                'amount': amount,
                'type': bill_type,
                'date': date_formatted,
                'category': category,
                'platform': 'alipay',
                'merchant': merchant[:100] if merchant else "",
                'note': get_val('note')[:200] if get_val('note') else "",
                'transaction_id': get_val('transaction_id')
            }
            
            bills.append(bill)
            
        except Exception as e:
            print(f"解析行失败: {row}, 错误: {e}")
            continue
    
    return bills

def parse_alipay_excel(content: bytes) -> List[dict]:
    bills = []
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    
    header_idx = None
    headers = []
    
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        row_values = [str(cell) if cell is not None else '' for cell in row]
        if '交易时间' in str(row_values) and ('交易分类' in str(row_values) or '交易对方' in str(row_values)):
            header_idx = row_idx
            headers = row_values
            break
    
    if header_idx is None:
        raise ValueError("无法识别支付宝账单格式：找不到表头")
    
    col_map = {}
    for i, h in enumerate(headers):
        h_str = str(h).strip()
        if '交易时间' in h_str:
            col_map['date'] = i
        elif '交易分类' in h_str or '交易类型' in h_str:
            col_map['type'] = i
        elif '交易对方' in h_str:
            col_map['merchant'] = i
        elif '商品说明' in h_str or '商品' in h_str:
            col_map['name'] = i
        elif '收/支' in h_str:
            col_map['direction'] = i
        elif '金额' in h_str:
            col_map['amount'] = i
        elif '收/付款方式' in h_str or '支付方式' in h_str:
            col_map['payment'] = i
        elif '交易订单号' in h_str or '交易号' in h_str:
            col_map['transaction_id'] = i
        elif '备注' in h_str:
            col_map['note'] = i
    
    for row in ws.iter_rows(min_row=header_idx + 1, values_only=True):
        row_values = [str(cell) if cell is not None else '' for cell in row]
        
        if not row_values or len(row_values) < 5:
            continue
        
        if '---' in row_values[0] or '以上是' in row_values[0]:
            break
        
        try:
            get_val = lambda key: row_values[col_map[key]].strip() if key in col_map and col_map[key] < len(row_values) else ""
            
            date_str = get_val('date')
            if not date_str:
                continue
            
            try:
                dt = datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
                date_formatted = dt.strftime("%Y-%m-%d")
            except:
                date_formatted = date_str[:10] if len(date_str) >= 10 else date_str
            
            direction = get_val('direction')
            bill_type = 'income' if '收入' in direction else 'expense' if '支出' in direction else 'expense'
            
            amount_str = get_val('amount')
            amount_str = re.sub(r'[¥￥,\s]', '', amount_str)
            try:
                amount = float(amount_str)
            except:
                continue
            
            if bill_type == 'expense':
                amount = -abs(amount)
            else:
                amount = abs(amount)
            
            merchant = get_val('merchant')
            name = get_val('name') or merchant
            
            category = categorize_transaction(name, get_val('type'))
            
            bill = {
                'name': name[:100] if name else "未知交易",
                'amount': amount,
                'type': bill_type,
                'date': date_formatted,
                'category': category,
                'platform': 'alipay',
                'merchant': merchant[:100] if merchant else "",
                'note': get_val('note')[:200] if get_val('note') else "",
                'transaction_id': get_val('transaction_id')
            }
            
            bills.append(bill)
            
        except Exception as e:
            print(f"解析行失败: {row_values}, 错误: {e}")
            continue
    
    return bills

def categorize_transaction(name: str, trans_type: str) -> str:
    name_lower = name.lower() if name else ""
    
    categories = {
        '餐饮': ['美团', '饿了么', '外卖', '餐厅', '食品', '星巴克', '肯德基', '麦当劳', '奶茶', '咖啡', '饭店', '小吃', '超市', '便利店'],
        '交通': ['滴滴', '打车', '地铁', '公交', '加油', '停车', '出行', '火车', '机票', '高铁'],
        '购物': ['淘宝', '京东', '拼多多', '天猫', '购物', '商城', '网购'],
        '娱乐': ['游戏', '视频', '音乐', '电影', '会员', '充值'],
        '医疗': ['医院', '药店', '医疗', '健康'],
        '教育': ['教育', '培训', '课程', '学习'],
        '转账': ['转账', '红包', '收款'],
        '工资': ['工资', '薪资', '代发'],
        '投资': ['理财', '基金', '股票', '收益'],
    }
    
    for category, keywords in categories.items():
        for keyword in keywords:
            if keyword in name_lower:
                return category
    
    if '转账' in trans_type or '红包' in trans_type:
        return '转账'
    
    return '其他'

@app.post("/bills/upload", response_model=ImportResult)
async def upload_bills(
    file: UploadFile = File(...),
    platform: str = Form(...)
):
    if platform not in ['wechat', 'alipay', 'bank']:
        raise HTTPException(status_code=400, detail="不支持的平台类型")
    
    if not file.filename.lower().endswith(('.csv', '.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="仅支持 CSV、Excel 格式文件")
    
    try:
        content = await file.read()
        
        is_excel = file.filename.lower().endswith(('.xlsx', '.xls'))
        
        if platform == 'wechat':
            if is_excel:
                bills = parse_wechat_excel(content)
            else:
                try:
                    content_str = content.decode('utf-8')
                except:
                    try:
                        content_str = content.decode('utf-8-sig')
                    except:
                        try:
                            content_str = content.decode('gbk')
                        except:
                            content_str = content.decode('gb18030')
                bills = parse_wechat_csv(content_str)
        elif platform == 'alipay':
            if is_excel:
                bills = parse_alipay_excel(content)
            else:
                try:
                    content_str = content.decode('utf-8')
                except:
                    try:
                        content_str = content.decode('utf-8-sig')
                    except:
                        try:
                            content_str = content.decode('gbk')
                        except:
                            content_str = content.decode('gb18030')
                bills = parse_alipay_csv(content_str)
        else:
            raise HTTPException(status_code=400, detail=f"{platform} 解析器暂未实现")
        
        if not bills:
            return ImportResult(
                success=0,
                skipped=0,
                total=0,
                message="未解析到有效账单数据"
            )
        
        existing_ids = set()
        all_bills = db.all()
        for b in all_bills:
            if b.get('transaction_id'):
                existing_ids.add(b.get('transaction_id'))
        
        success_count = 0
        skipped_count = 0
        
        for bill in bills:
            if bill.get('transaction_id') and bill['transaction_id'] in existing_ids:
                skipped_count += 1
                continue
            
            db.insert(bill)
            success_count += 1
        
        return ImportResult(
            success=success_count,
            skipped=skipped_count,
            total=len(bills),
            message=f"成功导入 {success_count} 条账单" + (f"，跳过 {skipped_count} 条重复记录" if skipped_count > 0 else "")
        )
        
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导入失败: {str(e)}")

@app.get("/bills/stats")
def get_stats():
    bills = db.all()
    
    total_income = sum(b.get('amount', 0) for b in bills if b.get('amount', 0) > 0)
    total_expense = sum(abs(b.get('amount', 0)) for b in bills if b.get('amount', 0) < 0)
    
    category_stats = {}
    for bill in bills:
        cat = bill.get('category', '其他')
        amount = abs(bill.get('amount', 0))
        if bill.get('amount', 0) < 0:
            category_stats[cat] = category_stats.get(cat, 0) + amount
    
    platform_stats = {}
    for bill in bills:
        plat = bill.get('platform', 'unknown')
        platform_stats[plat] = platform_stats.get(plat, 0) + 1
    
    return {
        "total_count": len(bills),
        "total_income": total_income,
        "total_expense": total_expense,
        "balance": total_income - total_expense,
        "category_stats": category_stats,
        "platform_stats": platform_stats
    }

@app.get("/")
def read_root():
    return {"message": "欢迎使用 BoboBill API! 访问 /docs 查看文档"}
